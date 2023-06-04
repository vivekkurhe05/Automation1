import openpyxl

class Common:

    def __init__(self, filepath, sheetname):
        global wk, sh
        self.filepath=filepath
        self.sheetname=sheetname
        wk = openpyxl.load_workbook(self.filepath)
        sh = wk[self.sheetname]

    def fetch_row_count(self):
        rows=sh.max_row
        return rows

    def fetch_column_count(self):
        cols=sh.max_column
        return cols

    def fetch_key_names(self):
        c=self.fetch_column_count()
        li=[]
        for i in range(1,c+1):
            cell=sh.cell(row=1, column=i)
            li.append(cell.value)
        return li

    def update_request_with_data(self,row_number,di,key_list):
        c=self.fetch_column_count()
        for i in range(1,c+1):
            cell=sh.cell(row=row_number,column=i)
            di[key_list[i-1]]=cell.value
        return di