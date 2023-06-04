# Data Driven Testing using Excel
import os.path

import requests
import json
import openpyxl

def test_add_multiple_students():
    URL="https://thetestingworldapi.com/api/studentsDetails"
    file=open(os.path.abspath(os.curdir)+"\\Data\\AddStudentDetails.json")
    di=json.loads(file.read())
    wk=openpyxl.load_workbook(os.path.abspath(os.curdir)+"\\test_data.xlsx")
    sh=wk['Sheet1']
    rows=sh.max_row
    for i in range(2, rows+1):
        first_name=sh.cell(row=i, column=1)
        middle_name=sh.cell(row=i, column=2)
        last_name=sh.cell(row=i, column=3)
        dob=sh.cell(row=i, column=4)
        di['first_name']=first_name.value
        di['middle_name']=middle_name.value
        di['last_name']=last_name.value
        di['date_of_birth']=dob.value
        response=requests.post(URL,di)
        print(response.text)
